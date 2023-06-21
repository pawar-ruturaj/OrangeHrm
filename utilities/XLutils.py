import openpyxl


def getrowCount(file, sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.max_row


def readData(file, sheetname, rowno, columnno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=rowno, column=columnno).value


def writeData(file, sheetname, rowno, columnno, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=rowno, column=columnno).value = data
    book.save(file)


